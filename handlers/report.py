import io
import logging
from telegram import Update
from telegram.ext import ContextTypes
from db.queries import get_user, get_entries_for_export
from handlers.summary import _resolve_period
from utils import fmt

logger = logging.getLogger(__name__)


async def send_report(reply_fn, reply_document_fn, user_row: dict, period_arg: str) -> None:
    try:
        start, end, label = _resolve_period(period_arg)
    except ValueError:
        await reply_fn(
            "Couldn't understand that period.\n"
            "Try: report week · report month · report last month · report 2 months · report year"
        )
        return

    # Respect history clear boundary
    cleared_at = user_row.get("history_cleared_at")
    if cleared_at:
        cleared_norm = str(cleared_at).replace(" ", "T")[:19]
        if start < cleared_norm:
            start = cleared_norm

    await reply_fn(f"Generating {label} report...")

    entries = get_entries_for_export(user_row["id"], start, end)
    if not entries:
        await reply_fn(f"No expenses found for {label}.")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        await reply_fn("openpyxl is not installed — cannot generate Excel file.")
        return

    currency = user_row["currency"]

    wb = openpyxl.Workbook()

    # ── Expenses sheet ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Expenses"

    headers = ["Date", "Category", "Item", "Amount", "Note"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    row_num = 2
    for entry in entries:
        if entry["items"]:
            for item in entry["items"]:
                ws.cell(row=row_num, column=1, value=entry["date"])
                ws.cell(row=row_num, column=2, value=entry["category"])
                ws.cell(row=row_num, column=3, value=item["name"])
                ws.cell(row=row_num, column=4, value=item["amount"])
                ws.cell(row=row_num, column=5, value=entry["raw_message"])
                row_num += 1
        else:
            ws.cell(row=row_num, column=1, value=entry["date"])
            ws.cell(row=row_num, column=2, value=entry["category"])
            ws.cell(row=row_num, column=3, value="")
            ws.cell(row=row_num, column=4, value=entry["total_amount"])
            ws.cell(row=row_num, column=5, value=entry["raw_message"])
            row_num += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = [12, 12, 20, 12, 30][col - 1]

    # ── Summary sheet ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Category").font = Font(bold=True)
    ws2.cell(row=1, column=2, value=f"Amount ({currency})").font = Font(bold=True)

    cat_totals: dict[str, float] = {}
    for entry in entries:
        cat_totals[entry["category"]] = cat_totals.get(entry["category"], 0.0) + entry["total_amount"]

    for i, (cat, total) in enumerate(
        sorted(cat_totals.items(), key=lambda x: x[1], reverse=True), start=2
    ):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=total)

    total_row = len(cat_totals) + 2
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=sum(cat_totals.values())).font = Font(bold=True)

    for col in (1, 2):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Send ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    grand_total = sum(cat_totals.values())
    safe_label  = label.lower().replace(" ", "_").replace("–", "-")
    filename    = f"spending_{safe_label}.xlsx"

    await reply_document_fn(
        document=buf,
        filename=filename,
        caption=f"{label} — {fmt(grand_total, currency)} total across {len(entries)} entries",
    )


async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_row = get_user(update.effective_user.id)
    if not user_row:
        await update.message.reply_text("Please run /start first.")
        return

    period_arg = " ".join(context.args) if context.args else "month"
    await send_report(
        update.message.reply_text,
        update.message.reply_document,
        user_row,
        period_arg,
    )
